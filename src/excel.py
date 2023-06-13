import logging
import os.path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime


def load_wb(filename):
    # Make sure the excel sheet is not open
    i = 0
    while True:
        if i > 5:
            raise RuntimeError("You tried too many times")
        try:
            f = open(filename, 'r+')
            break
        except IOError:
            input("Please close the open file in Excel and press enter. ")
            logging.warning("User tried to open a file that is locked and is likely open in Excel.")
            i += 1

    return load_workbook(filename)


def save_wb(filename, wb):
    try:
        wb.save(filename)
    except PermissionError as E:
        logging.warning("Permission error when saving.\n{}".format(E))
        filename += "Conflicted Copy.xlsx"
        wb.save(filename)


def next_table_name(wb):
    tables = {}
    for sheet in wb.sheetnames:
        tables = tables | wb[sheet].tables
    base = 1
    while True:
        if base > 100:
            raise RuntimeError("Too many tables")
        if "Table{}".format(base) in tables:
            base += 1
        else:
            return "Table{}".format(base)


def dataframe(form_fields, form_type):
    return form_fields[form_type]["dataframe"]


def append_table(wb, data, form_fields, form_kind, worksheet=None):
    if worksheet.lower()[-9:] == ' (legacy)':
        worksheet = worksheet[:-9]
    if worksheet is None:
        ws = wb.active
    else:
        if worksheet in wb:
            ws = wb[worksheet]
        else:
            logging.error("Could not find a worksheet named \"{}\" in the workbook. Please try to create it.")
            raise IndexError("Worksheet not found in workbook.")
    result = []
    for i in dataframe(form_fields, form_kind).keys():
        result.append(data[i])
    ws.append(result)
    return wb


def new_wb():
    wb = Workbook()
    return wb


def filename_generate(dataframe):
    name = "".join(dataframe["Name"].split(" "))
    date = "-".join(dataframe["Checkout End"].split("/"))
    return "EquipmentChk_{}_{}.pdf".format(name, date)
